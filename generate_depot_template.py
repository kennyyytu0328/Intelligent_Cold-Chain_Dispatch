"""
生成倉庫資料 Excel 範本檔案

使用方式:
    python generate_depot_template.py

會生成 ICCDDS_Depot_Template.xlsx 檔案
"""
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 設置 Windows 控制台編碼
if sys.platform == 'win32':
    try:
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    except:
        pass


def create_depot_template():
    """創建倉庫資料 Excel 範本"""

    # ===== 示範資料 =====
    depots_data = [
        {
            "name": "台北101倉庫",
            "code": "TP-001",
            "address": "台北市信義區信義路五段7號",
            "latitude": 25.0330,
            "longitude": 121.5654,
            "is_active": True,
            "contact_person": "王大明",
            "contact_phone": "02-8101-8800",
        },
        {
            "name": "板橋物流中心",
            "code": "BD-001",
            "address": "新北市板橋區縣民大道二段7號",
            "latitude": 25.0122,
            "longitude": 121.4627,
            "is_active": True,
            "contact_person": "李小華",
            "contact_phone": "02-2960-3456",
        },
        {
            "name": "內湖配送站",
            "code": "NH-001",
            "address": "台北市內湖區民權東路六段",
            "latitude": "",  # 空白，需要轉換
            "longitude": "",  # 空白，需要轉換
            "is_active": True,
            "contact_person": "張三",
            "contact_phone": "02-2792-8888",
        },
    ]

    # ===== 使用說明 =====
    instructions = [
        ["欄位名稱", "說明", "必填", "範例"],
        ["name", "倉庫名稱", "是", "台北101倉庫"],
        ["code", "倉庫代碼（唯一）", "否", "TP-001"],
        ["address", "完整地址", "否*", "台北市信義區信義路五段7號"],
        ["latitude", "緯度（小數）", "否*", "25.0330"],
        ["longitude", "經度（小數）", "否*", "121.5654"],
        ["is_active", "是否啟用", "是", "TRUE 或 FALSE"],
        ["contact_person", "聯絡人", "否", "王大明"],
        ["contact_phone", "聯絡電話", "否", "02-8101-8800"],
        ["", "", "", ""],
        ["重要說明", "", "", ""],
        ["1. 座標輸入方式", "方式 A：填寫 latitude 和 longitude → 直接使用", "", ""],
        ["", "方式 B：只填寫 address → 系統自動轉換（需要 1 秒/筆）", "", ""],
        ["2. 批量匯入建議", "如果有 10 筆以上資料，建議先在 Excel 填好經緯度", "", ""],
        ["", "可使用 Google Maps 查詢座標：右鍵點擊 → 複製座標", "", ""],
        ["3. 速率限制", "使用地址轉換時，每秒最多轉換 1 筆", "", ""],
        ["", "例如：50 筆資料約需 50 秒", "", ""],
        ["4. is_active 欄位", "TRUE = 啟用，FALSE = 停用", "", ""],
        ["5. code 唯一性", "倉庫代碼不可重複", "", ""],
    ]

    # ===== 建立 Excel 檔案 =====
    filename = "ICCDDS_Depot_Template.xlsx"

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Sheet 1: 使用說明
        pd.DataFrame(instructions).to_excel(
            writer,
            sheet_name='使用說明',
            index=False,
            header=False
        )

        # Sheet 2: 倉庫資料
        pd.DataFrame(depots_data).to_excel(
            writer,
            sheet_name='倉庫 (Depots)',
            index=False
        )

    # ===== 美化格式 =====
    wb = load_workbook(filename)

    # 格式化使用說明 Sheet
    ws_instructions = wb['使用說明']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws_instructions[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 設置欄寬
    ws_instructions.column_dimensions['A'].width = 25
    ws_instructions.column_dimensions['B'].width = 50
    ws_instructions.column_dimensions['C'].width = 10
    ws_instructions.column_dimensions['D'].width = 30

    # 格式化倉庫資料 Sheet
    ws_depots = wb['倉庫 (Depots)']

    for cell in ws_depots[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 設置欄寬
    ws_depots.column_dimensions['A'].width = 20  # name
    ws_depots.column_dimensions['B'].width = 12  # code
    ws_depots.column_dimensions['C'].width = 35  # address
    ws_depots.column_dimensions['D'].width = 12  # latitude
    ws_depots.column_dimensions['E'].width = 12  # longitude
    ws_depots.column_dimensions['F'].width = 12  # is_active
    ws_depots.column_dimensions['G'].width = 15  # contact_person
    ws_depots.column_dimensions['H'].width = 15  # contact_phone

    wb.save(filename)

    return filename


def main():
    print("📋 正在生成倉庫 Excel 範本...")
    print()

    filename = create_depot_template()

    print(f"✅ 範本已生成: {filename}")
    print()
    print("📝 檔案包含:")
    print("   • 使用說明 (Instructions)")
    print("   • 倉庫資料範例 (3 筆)")
    print()
    print("🎯 匯入方式:")
    print("   1. 編輯 Excel 檔案，填寫倉庫資料")
    print("   2. 座標填寫方式二選一:")
    print("      A. 填寫 latitude 和 longitude")
    print("      B. 只填寫 address（系統自動轉換，約 1 秒/筆）")
    print("   3. 透過前端 UI 上傳檔案匯入")
    print()
    print("⚠️  注意事項:")
    print("   • 建議批量匯入時預先填好座標（避免等待過久）")
    print("   • 倉庫代碼 (code) 不可重複")
    print("   • is_active 欄位請填 TRUE 或 FALSE")
    print()


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"❌ 發生錯誤: {e}")
        import traceback
        traceback.print_exc()
